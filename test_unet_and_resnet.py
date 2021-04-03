import tensorflow as tf
from keras import backend as K
from keras.models import load_model
from scipy.ndimage import label
import numpy as np

from misc.utils import pre, rec, F, grayscale_to_rgb, dice_coef, dice_coef_loss, window_setting_for_ct, MyClass
from generator.generator_for_3d_resnet import BonePatchSequence
from misc.save_images_and_raw import save_images, save_raw
from train_3d_unet import dc, se, focal_tversky_loss_08

import openpyxl
import time
import argparse
import datetime
import pytz
import pprint
import os


def set_arrays(args, pt):

    a = MyClass()

    a.ct = np.load(args.test_images_dir + pt + "_ct.npy")
    a.lb = np.load(args.test_images_dir + pt + "_lb.npy")
    a.bs = np.load(args.test_images_dir + pt + "_bs.npy")
    if a.ct.ndim == 3:
        a.ct = a.ct[:, :, :, np.newaxis]
        a.lb = a.lb[:, :, :, np.newaxis]
        a.bs = a.bs[:, :, :, np.newaxis]
    a.bbs = np.zeros_like(a.ct)

    a.lb_0or1 = np.where((a.lb > 0) & (a.lb < 1000), 1, 0)
    a.ct_rgb = grayscale_to_rgb(window_setting_for_ct(a.ct))
    a.ct_rgb_with_pred = np.copy(a.ct_rgb)
    a.ct_rgb_with_window = np.copy(a.ct_rgb)

    a.width = a.ct.shape[1]
    a.height = a.ct.shape[0]

    a.margin_xy = (a.width % 16) // 2
    a.margin_z = (a.height % 16) // 2

    return a


def run_unet_and_resnet(args):

    config = tf.ConfigProto()
    config.gpu_options.per_process_gpu_memory_fraction = 0.8
    sess = tf.Session(config=config)
    K.set_session(sess)

    np.set_printoptions(precision=3, suppress=True, linewidth=np.inf)

    V = 100

    pt_list = os.listdir(args.test_images_dir)

    ####
    pt_list = [x[:5] for x in pt_list]
    pt_list = list(set(pt_list))
    ####

    pt_list.sort()

    start_test = time.time()

    print("### Test on iso-arrays. ###")
    print(args.test_images_dir)
    print(pt_list)
    print(len(pt_list))

    TP = 0
    FN = 0
    FP = 0

    sheet0 = [["ReaderID", "ModalityID", "CaseID", "LesionID", "TP_Rating", "DSC"]]
    sheet1 = [["ReaderID", "ModalityID", "CaseID", "FP_Rating"]]
    sheet2 = [["CaseID", "LesionID", "Weight", "ReaderID", "ModalityID"]]

    for n in range(len(pt_list)):

        start_pt = time.time()

        pt_name = pt_list[n]
        print("#", pt_name)

        a = set_arrays(args, pt_name)  # set instance to store arrays.
        a.bbs_96 = np.copy(a.bbs)
        a.seg = np.copy(a.bbs).astype("float32")

        lesion_num_list = np.unique(a.lb).tolist()
        del lesion_num_list[0]

        
        ################
        ### Run UNet ###
        ################
        
        start_unet = time.time()

        a.ct_expanded = np.ones((a.ct.shape[0] + 48, a.ct.shape[1] + 48, a.ct.shape[2] + 48, 1)) * (-1000)
        a.ct_expanded[24:a.ct.shape[0] + 24, 24:a.ct.shape[1] + 24, 24:a.ct.shape[2] + 24, 0] = a.ct[:, :, :, 0]
        a.bs_expanded = np.zeros((a.bs.shape[0] + 48, a.bs.shape[1] + 48, a.bs.shape[2] + 48, 1))
        a.bs_expanded[24:a.bs.shape[0] + 24, 24:a.bs.shape[1] + 24, 24:a.bs.shape[2] + 24, 0] = a.bs[:, :, :, 0]
        a.ct_copy = np.copy(a.ct)
        a.bs_copy = np.copy(a.bs)
        a.ct = a.ct_expanded
        a.bs = a.bs_expanded

        a.bbs_96_expanded = np.zeros_like(a.ct_expanded)
        a.seg_expanded = np.zeros_like(a.ct_expanded, "float32")

        test_gen = BonePatchSequence(a, L=96, ws=[2000, 400, 600, 150], with_coordinate=True)

        model = load_model(args.unet_model, custom_objects={"dice_coef_loss": dice_coef_loss, "dice_coef": dice_coef,
                                                            "focal_tversky_loss_08": focal_tversky_loss_08, "dc": dc, "se": se})

        for X, c in test_gen:
            x, y, z = c
            pred = model.predict(X)
            a.bbs_96_expanded[z:z + 96, y:y + 96, x:x + 96, :] += 1
            a.seg_expanded[z:z + 96, y:y + 96, x:x + 96, :] += pred[0]

        K.clear_session()

        a.seg = a.seg_expanded[24:-24, 24:-24, 24:-24, :]
        a.bbs_96 = a.bbs_96_expanded[24:-24, 24:-24, 24:-24, :]
        a.seg_denom = np.where(a.bbs_96 == 0, 1, a.bbs_96)
        a.seg = a.seg / a.seg_denom
        a.seg_0or1 = np.where(a.seg > 0.1, 1, 0)

        a.ct = a.ct_copy
        a.bs = a.bs_copy

        unet_time = time.time() - start_unet


        ################
        ### Trimming ###
        ################

        start_trimming = time.time()

        a.seg_labeled, num_features = label(a.seg_0or1)
        a.seg_0or1 = np.zeros_like(a.seg_0or1)

        for i in range(num_features):
            lb = (a.seg_labeled == i + 1)
            volume = np.sum(lb)
            if volume > V:
                a.seg_0or1 += np.where(a.seg_labeled == i + 1, 1, 0)

        a.seg_labeled, num_features = label(a.seg_0or1)

        label_info_array = np.zeros((num_features, 6))      # [id, volume, flag1, flag2, flag3, FPorTP]

        for i in range(num_features):
            lb = (a.seg_labeled == i + 1)
            volume = np.sum(lb)

            label_info_array[i][0] = i + 1
            label_info_array[i][1] = volume

        trimming_time = time.time() - start_trimming


        ##################################
        ### Run VGG or ResNet ensemble ###
        ##################################

        start_vgg = time.time()

        test_gen = BonePatchSequence(a, L=32, B=1, ws=[2000, 400, 600, 150], with_coordinate=True)


        print("calculation for 1st model, start.")
        model = load_model(args.resnet_model[0], custom_objects={"pre": pre, "rec": rec, "F": F})
        for X, c in test_gen:
            x, y, z = c
            if np.sum(a.seg_0or1[z:z + 32, y:y + 32, x:x + 32, :]) > V:
                a.bbs[z:z + 32, y:y + 32, x:x + 32, :] += 1
                pred = model.predict(X)
                label_nums = np.unique(a.seg_labeled[z:z + 32, y:y + 32, x:x + 32, :]).tolist()
                label_nums.remove(0)
                for j in label_nums:
                    label_info_array[j-1][2] = np.max([label_info_array[j-1][2], pred[:, 1]])
        K.clear_session()


        if len(args.resnet_model) >= 2 and os.path.exists(args.resnet_model[1]):
            print("calculation for 2nd model, start.")
            model = load_model(args.resnet_model[1], custom_objects={"pre": pre, "rec": rec, "F": F})
            for X, c in test_gen:
                x, y, z = c
                if np.sum(a.seg_0or1[z:z + 32, y:y + 32, x:x + 32, :]) > V:
                    pred = model.predict(X)
                    label_nums = np.unique(a.seg_labeled[z:z + 32, y:y + 32, x:x + 32, :]).tolist()
                    label_nums.remove(0)
                    for j in label_nums:
                        label_info_array[j - 1][3] = np.max([label_info_array[j - 1][3], pred[:, 1]])
            K.clear_session()


        if len(args.resnet_model) >= 3 and os.path.exists(args.resnet_model[2]):
            print("calculation for 3rd model, start.")
            model = load_model(args.resnet_model[2], custom_objects={"pre": pre, "rec": rec, "F": F})
            for X, c in test_gen:
                x, y, z = c
                if np.sum(a.seg_0or1[z:z + 32, y:y + 32, x:x + 32, :]) > V:
                    pred = model.predict(X)
                    label_nums = np.unique(a.seg_labeled[z:z + 32, y:y + 32, x:x + 32, :]).tolist()
                    label_nums.remove(0)
                    for j in label_nums:
                        label_info_array[j - 1][4] = np.max([label_info_array[j - 1][4], pred[:, 1]])
            K.clear_session()

        print(label_info_array)

        a.seg_0or1 = np.zeros_like(a.seg_0or1)
        label_info_list = []
        for i in range(label_info_array.shape[0]):
            conf = np.max(label_info_array[i][2:5])
            if conf >= args.thr:
                a.seg_0or1 += np.where(a.seg_labeled==label_info_array[i][0], 1, 0)
            if conf >= 0.1:
                label_info_array[i][5] = conf
                label_info_list.append(label_info_array[i].tolist())

        for line in label_info_list:
            line = [round(i, 3) for i in line]
            print(line)

        vgg_time = time.time() - start_vgg


        ####################
        ### measurements ###
        ####################

        start_measurement = time.time()

        lesion_num_list = np.unique(a.lb).tolist()
        del lesion_num_list[0]

        labeled_array, num_regions = label(a.seg_0or1)

        dsc_list = ""

        for i in lesion_num_list:

            a.lesion_i_lb = np.where(a.lb == i, 1, 0)

            overlap = (a.lesion_i_lb * a.seg_0or1).sum()

            overlap_regions_sum = 0
            for j in range(num_regions):
                a.region_j_lb = np.where(labeled_array == j + 1, 1, 0)
                if (a.lesion_i_lb * a.region_j_lb).sum() > 0:
                    overlap_regions_sum += a.region_j_lb.sum()

            dsc = overlap * 2 / (a.lesion_i_lb.sum() + overlap_regions_sum)
            dsc_list += str(round(dsc, 2)) + " "

            if dsc > args.dsc_thr:
                TP += 1
            if dsc == 0:
                FN += 1

        print(dsc_list)


        for i in range(num_regions):
            if (a.lb * np.where(labeled_array == i + 1, 1, 0)).sum() == 0:
                FP += 1

        print(TP, FN, FP)


        #######################
        ### Add Excel Sheet ###
        #######################

        overlap_ids_all = []

        for ii, i in enumerate(lesion_num_list):
            line2 = [int(pt_name[2:]), ii+1, 0, 0, 0]
            sheet2.append(line2)

            a.lesion_i_lb = np.where(a.lb == i, 1, 0)
            a.overlap = a.seg_labeled * a.lesion_i_lb
            overlap_ids = np.unique(a.overlap).tolist()
            del overlap_ids[0]

            overlap_volume = np.sum(a.seg_0or1 * a.lesion_i_lb)
            gt_volume = np.sum(a.lesion_i_lb)
            pred_volume = 0

            conf_level = 0

            for j in range(len(label_info_list)):
                if label_info_list[j][0] in overlap_ids:
                    conf_level = max([label_info_list[j][5], conf_level])
                    if label_info_list[j][5] >= args.thr:
                        pred_volume += np.sum(np.where(a.seg_labeled == label_info_list[j][0], 1, 0))

            dsc = 2 * overlap_volume / (gt_volume + pred_volume + 1)
            print(i, j, gt_volume, pred_volume, dsc)

            if conf_level > 0.1:
                line0 = [0, 0, int(pt_name[2:]), ii+1, round(conf_level, 5), dsc]
                sheet0.append(line0)

            overlap_ids_all += overlap_ids

        if len(lesion_num_list) == 0:
            line2 = [int(pt_name[2:]), 0, 0, 0, 0]
            sheet2.append(line2)


        for j in range(len(label_info_list)):
            if label_info_list[j][0] not in overlap_ids_all:
                line1 = [0, 0, int(pt_name[2:]), round(label_info_list[j][5], 5)]

                sheet1.append(line1)

        pprint.pprint(sheet0)
        pprint.pprint(sheet1)
        pprint.pprint(sheet2)

        measurement_time = time.time() - start_measurement


        ###################
        ### Save images ###
        ###################

        start_save_images = time.time()

        if args.si == 1:
            save_images(args, pt_name, a, label_info_array)

        save_images_time = time.time() - start_save_images


        ################
        ### Save raw ###
        ################

        start_save_raw = time.time()

        if args.sr == 1:
            save_raw(args, pt_name, a)

        save_raw_time = time.time() - start_save_raw


        pt_time = time.time() - start_pt

        print("#", pt_name, "unet", round(unet_time, 3), "trimming", round(trimming_time, 3), "resnet", round(vgg_time, 3), "measurement",
              round(measurement_time, 3), "save images", round(save_images_time, 3), "save raw", round(save_raw_time, 3), "pt total", round(pt_time, 3))



    #############################
    ### Save Config and Excel ###
    #############################

    text_path = args.results_dir + args.dt + "_config.txt"
    xlsx_path = args.results_dir + args.dt + "_JAFROC_Datafile.xlsx"

    with open(text_path, mode="w") as f:
        f.write("res_thr" + str(args.res_thr) + "\n")
        f.write(os.path.basename(args.resnet_model) + "\n")
        f.write(os.path.basename(args.resnet_model_2) + "\n")
        f.write(os.path.basename(args.resnet_model_3) + "\n")
        f.write("\n")
        f.write(os.path.basename(args.unet_model) + "\n")
        f.write("\n")
        f.write("V" + str(V) + "\n")
        f.write(str(args.memo) + "\n")
        f.write("\n")
        f.write(str(TP) + " " + str(FN) + " " + str(FP) + "\n")

    wb = openpyxl.Workbook()
    for (sheet, sheet_name) in [(sheet0, "TP"), (sheet1, "FP"), (sheet2, "TRUTH")]:
        wb.create_sheet(sheet_name)
        wb.active = wb.sheetnames.index(sheet_name)
        ws = wb.active
        for line in sheet:
            ws.append(line)
    wb.save(xlsx_path)

    test_time = time.time() - start_test
    print("total test time :", test_time)

    return


####################
## ArgumentParser ##
####################

def set_args():

    parser = argparse.ArgumentParser(argument_default=argparse.SUPPRESS)

    parser.add_argument("--dicom_dir", type=str, default="/xxx/dicom_dir")
    parser.add_argument("--test_images_dir", type=str, default="/xxx/dir3")
    parser.add_argument("--results_dir", type=str, default="/xxx/results")

    parser.add_argument("--unet_model", type=str, default="./unet.h5")
    parser.add_argument("--resnet_model", type=str, nargs="*", default=["./resnet1.h5",
                                                                        "./resnet2.h5",
                                                                        "./resnet3.h5"])

    parser.add_argument("--thr", type=float, default=0.6)
    parser.add_argument("--dsc_thr", type=float, default=0.3)

    parser.add_argument("--si", type=int, default=1, help="save predicted images.")
    parser.add_argument("--sr", type=int, default=0, help="save raw files of predicted labels.")

    Args = parser.parse_args()

    assert os.path.exists(Args.unet_model)
    assert os.path.exists(Args.resnet_model[0])

    Args.dt = datetime.datetime.now(pytz.timezone("Asia/Tokyo")).strftime('%Y%m%d_%H%M%S')

    Args.results_dir = Args.results_dir + Args.dt + "/"
    os.makedirs(Args.results_dir)

    return Args


##########
## Main ##
##########

if __name__ == '__main__':

    Args = set_args()
    run_unet_and_resnet(Args)
